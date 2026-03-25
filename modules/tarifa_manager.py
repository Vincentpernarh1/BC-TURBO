"""
Módulo para gerenciamento de dados de Tarifa por Fluxo
Processa arquivos Excel de diferentes tipos de fluxo e gera parquets para performance
"""

import pandas as pd
import numpy as np
import os
import re
import unicodedata
import openpyxl
from pathlib import Path


class TarifaManager:
    def __init__(self, db_folder=None):
        self.db_folder = db_folder
        self.tarifa_base_folder = None
        self.fluxo_data = {}  # Dict: {fluxo_name: DataFrame}
        
    def _parse_transporter_name(self, filename):
        """Extrai nome da transportadora do nome do arquivo"""
        match = re.search(r'_(.*?)\.', filename)
        if match:
            return match.group(1).replace('_', ' ').title()
        return os.path.splitext(filename)[0].replace('_', ' ').title()
    
    def _normalize_vehicle_name(self, name):
        """Normaliza nomes de veículos para padronização"""
        if not name or not str(name).strip():
            return None
        clean_name = str(name).strip().upper()
        if 'BITREM' in clean_name:
            return 'BITREM'
        if 'VANDERLEIA' in clean_name:
            return 'VANDERLEIA'
        if 'CARRETA' in clean_name:
            return 'CARRETA'
        if 'VAN' in clean_name or 'DUCATO' in clean_name:
            return 'VAN'
        if '3/4' in clean_name or '0.75' in clean_name:
            return '3/4'
        if 'TOCO' in clean_name:
            return 'TOCO'
        if 'TRUCK' in clean_name:
            return 'TRUCK'
        if 'FIORINO' in clean_name:
            return 'FIORINO'
        return clean_name
    
    def _normalize_text(self, text):
        """Remove acentos e normaliza texto"""
        if isinstance(text, str):
            # Remove accents/diacritics
            text = ''.join(
                c for c in unicodedata.normalize('NFKD', text)
                if not unicodedata.combining(c)
            )
            return text.strip().title()
        return text
    
    def _needs_parquet_conversion(self, source_folder, parquet_path):
        """Verifica se os arquivos Excel precisam ser convertidos para Parquet"""
        parquet_file = Path(parquet_path)
        
        # Se Parquet não existe, precisa converter
        if not parquet_file.exists():
            return True
        
        # Verifica se algum Excel na pasta foi modificado depois do Parquet
        parquet_mtime = parquet_file.stat().st_mtime
        
        for file in Path(source_folder).glob("*.xlsx"):
            if '~$' not in file.name and file.stat().st_mtime > parquet_mtime:
                return True
        
        for file in Path(source_folder).glob("*.xls"):
            if '~$' not in file.name and file.stat().st_mtime > parquet_mtime:
                return True
        
        return False
    
    def _process_milk_run_fluxo(self, fluxo_path, fluxo_name):
        """Processa arquivos do tipo MILK RUN"""
        all_melted_dfs = []
        
        for file_name in os.listdir(fluxo_path):
            if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~$'):
                continue
            
            file_path = os.path.join(fluxo_path, file_name)
            try:
                full_df = pd.read_excel(file_path, header=None, engine='openpyxl', keep_default_na=False, na_values=[''])

                header_row_idx = -1
                faixa_col_idx = -1
                
                for r_idx in range(min(10, len(full_df))):
                    row_as_str = full_df.iloc[r_idx].astype(str).str.strip().str.upper()
                    matches = row_as_str[row_as_str == 'FAIXA KM']
                    if not matches.empty:
                        header_row_idx = r_idx
                        faixa_col_idx = matches.index[0]
                        break
                
                if header_row_idx == -1:
                    print(f"  Skipping file {file_name}: Could not find 'FAIXA KM' header.")
                    continue

                rt_info_row_idx = header_row_idx - 1
                data_start_row_idx = header_row_idx + 1

                if rt_info_row_idx < 0 or data_start_row_idx >= len(full_df):
                    print(f"  Skipping file {file_name}: Invalid structure around 'FAIXA KM' header.")
                    continue

                # Carry over the trip type to all vehicles
                round_trip_info = full_df.iloc[rt_info_row_idx].ffill()
                vehicle_headers = full_df.iloc[header_row_idx]
                data_df = full_df.iloc[data_start_row_idx:]
                
                vehicle_col_indices = [
                    idx for idx, val in enumerate(vehicle_headers)
                    if idx > faixa_col_idx and pd.notna(val) and str(val).strip() != ""
                ]

                processed_rows = []
                for _, data_row in data_df.iterrows():
                    faixa_km_str = str(data_row.iloc[faixa_col_idx]).strip()
                    distancia_min = distancia_max = None

                    match = re.search(r'(\d+)\s*(?:a|-|até)\s*(\d+)', faixa_km_str, re.IGNORECASE)
                    if match:
                        distancia_min, distancia_max = int(match.group(1)), int(match.group(2))
                    else:
                        match = re.search(r'acima de\s*(\d+)', faixa_km_str, re.IGNORECASE)
                        if match:
                            distancia_min, distancia_max = int(match.group(1)), float(200)
                        else:
                            match = re.search(r'até\s*(\d+)', faixa_km_str, re.IGNORECASE)
                            if match:
                                distancia_min, distancia_max = 0, int(match.group(1))

                    if distancia_min is None:
                        continue

                    for col_idx in vehicle_col_indices:
                        try:
                            vehicle = vehicle_headers.iloc[col_idx]
                            viagem = round_trip_info.iloc[col_idx]
                            tarifa = data_row.iloc[col_idx]

                            if pd.notna(tarifa) and str(tarifa).strip() not in ("", "nan") and pd.notna(vehicle) and pd.notna(viagem):
                                viagem_str = str(viagem).upper().strip()
                                if 'ROUND' in viagem_str:
                                    viagem_code = 'RT'
                                elif 'ONE WAY' in viagem_str or 'OW' in viagem_str:
                                    viagem_code = 'OW'
                                else:
                                    viagem_code = viagem_str
                                
                                processed_rows.append({
                                    'Nomeacao': 'N/A',
                                    'Fornecedor': 'N/A',
                                    'Origem': 'N/A',
                                    'LocalColeta': 'N/A',
                                    'Destino': 'N/A',
                                    'DistanciaMin': distancia_min,
                                    'DistanciaMax': distancia_max,
                                    'Transportadora': self._parse_transporter_name(file_name),
                                    'Veiculo': str(vehicle),
                                    'Viagem': viagem_code,
                                    'Tarifa': tarifa,
                                    'Chave': 'N/A & N/A'
                                })
                        except Exception as inner_e:
                            print(f"  Error processing vehicle column in {file_name}: {inner_e}")
                            continue
                    
                if processed_rows:
                    all_melted_dfs.append(pd.DataFrame(processed_rows))

            except Exception as e:
                print(f"  Error processing file {file_path} for 'MILK RUN': {e}")
                continue
        
        return all_melted_dfs
    
    def _process_faixa_fluxo(self, fluxo_path, fluxo_name):
        """Processa arquivos do tipo FAIXA"""
        all_melted_dfs = []
        
        for file_name in os.listdir(fluxo_path):
            if not file_name.lower().endswith(('.xlsx', '.xls')):
                continue
            file_path = os.path.join(fluxo_path, file_name)
            try:
                full_df = pd.read_excel(file_path, header=None, engine='openpyxl', keep_default_na=False, na_values=[''])

                vehicles = full_df.iloc[0].ffill()
                headers = full_df.iloc[1]

                # Start reading from row 4 (index 3)
                data_df = full_df.iloc[3:].copy()
                data_df.columns = headers

                rename_map = {}
                for col in data_df.columns:
                    col_str = str(col).lower().strip()
                    if col_str == 'origem':
                        rename_map[col] = 'Origem'
                    elif col_str == 'destino':
                        rename_map[col] = 'Destino'
                data_df.rename(columns=rename_map, inplace=True)

                origem_indices = [i for i, col in enumerate(data_df.columns) if col == 'Origem']
                destino_indices = [i for i, col in enumerate(data_df.columns) if col == 'Destino']

                if not origem_indices or not destino_indices:
                    print(f"  Skipping file {file_name} due to missing 'Origem' or 'Destino' columns.")
                    continue

                origem_idx = origem_indices[0]
                destino_idx = destino_indices[0]

                processed_rows = []
                for _, data_row in data_df.iterrows():
                    origem = data_row.iloc[origem_idx]
                    destino_full = data_row.iloc[destino_idx]

                    destino = destino_full
                    distancia_min = None
                    distancia_max = None

                    if isinstance(destino_full, str):
                        # Extract prefix (UF code like "SP", "MG", etc.)
                        destino = destino_full[:2].strip()

                        # Extract "de X a Y" using regex
                        match = re.search(r'de\s*(\d+)\s*a\s*(\d+)', destino_full, flags=re.IGNORECASE)
                        if match:
                            try:
                                distancia_min = int(match.group(1))
                                distancia_max = int(match.group(2))
                            except Exception as e:
                                print(f"  Erro ao converter Distancia em '{destino_full}': {e}")

                    for col_idx, _ in enumerate(data_df.columns):
                        if col_idx == origem_idx or col_idx == destino_idx:
                            continue

                        vehicle = vehicles.iloc[col_idx]
                        viagem = headers.iloc[col_idx]
                        tarifa = data_row.iloc[col_idx]

                        if pd.notna(tarifa) and str(tarifa).strip() != "" and pd.notna(vehicle) and pd.notna(viagem):
                            processed_rows.append({
                                'Nomeacao': 'N/A',
                                'Fornecedor': 'N/A',
                                'Origem': origem,
                                'LocalColeta': 'N/A',
                                'Destino': destino,
                                'DistanciaMin': distancia_min,
                                'DistanciaMax': distancia_max,
                                'Transportadora': self._parse_transporter_name(file_name),
                                'Veiculo': str(vehicle),
                                'Viagem': str(viagem),
                                'Tarifa': tarifa,
                                'Chave': str(origem) + ' & ' + str(destino)
                            })

                if processed_rows:
                    all_melted_dfs.append(pd.DataFrame(processed_rows))

            except Exception as e:
                print(f"  Error processing file {file_path} for 'FAIXA': {e}")
                continue
        
        return all_melted_dfs
    
    def _process_spots_fluxo(self, fluxo_path, fluxo_name):
        """Processa arquivos do tipo SPOTS"""
        all_melted_dfs = []
        
        for file_name in os.listdir(fluxo_path):
            if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~$'):
                continue
            file_path = os.path.join(fluxo_path, file_name)
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb.active
                motorista_cols = {}
                last_vehicle = None
                
                for col_idx in range(1, sheet.max_column + 1):
                    cell_val = sheet.cell(row=1, column=col_idx).value
                    if cell_val and str(cell_val).strip():
                        vehicle_name = str(cell_val).strip()
                        if vehicle_name == '0.75':
                            last_vehicle = '3/4'
                        else:
                            last_vehicle = vehicle_name
                    
                    if last_vehicle:
                        motorista_val_raw = sheet.cell(row=3, column=col_idx).value
                        if motorista_val_raw is not None:
                            try:
                                motorista_clean = int(float(str(motorista_val_raw).strip()))
                                motorista_cols[col_idx] = (last_vehicle, motorista_clean)
                            except (ValueError, TypeError):
                                continue

                header_map = {}
                data_start_row, header_found_row = 1, -1
                for r in range(1, min(10, sheet.max_row + 1)):
                    if header_found_row != -1 and r > header_found_row:
                        break
                    for c in range(1, min(20, sheet.max_column + 1)):
                        cell_val = str(sheet.cell(row=r, column=c).value or '').strip().lower()
                        if 'origem' in cell_val:
                            header_map['Origem'] = c
                        elif 'destino' in cell_val:
                            header_map['Destino'] = c
                    if 'Origem' in header_map or 'Destino' in header_map:
                        header_found_row, data_start_row = r, r + 1
                
                if 'Origem' not in header_map or 'Destino' not in header_map:
                    continue
                
                processed_rows = []
                for row_idx in range(data_start_row, sheet.max_row + 1):
                    origem = str(sheet.cell(row=row_idx, column=header_map['Origem']).value or '').strip()
                    destino_raw = str(sheet.cell(row=row_idx, column=header_map['Destino']).value or '').strip()
                    if not origem or not destino_raw:
                        continue

                    # Parse distance and destination
                    distancia_min, distancia_max = None, None
                    clean_destino = destino_raw

                    # Scenario 1: "PE 01 KM - 10 Km" or "MG 11-20"
                    match = re.search(r'^(.*?)\s*(\d+)\s*(?:km)?\s*-\s*(\d+)', destino_raw, re.IGNORECASE)
                    if match:
                        clean_destino = match.group(1).strip()
                        distancia_min, distancia_max = int(match.group(2)), int(match.group(3))
                    else:
                        # Scenario 2: "De 21 km a 30 km" -> Destino becomes the same as Origem
                        match = re.search(r'de\s*(\d+)\s*(?:a|-|até)\s*(\d+)', destino_raw, re.IGNORECASE)
                        if match:
                            distancia_min, distancia_max = int(match.group(1)), int(match.group(2))
                            clean_destino = origem
                        else:
                            # Scenario 3: "BA acima 40 km"
                            match = re.search(r'^(.*?)\s*acima (?:de)?\s*(\d+)', destino_raw, re.IGNORECASE)
                            if match:
                                clean_destino = match.group(1).strip()
                                distancia_min, distancia_max = int(match.group(2)), float(200)
                            else:
                                # Scenario 4: "SE até 40 km"
                                match = re.search(r'^(.*?)\s*até\s*(\d+)', destino_raw, re.IGNORECASE)
                                if match:
                                    clean_destino = match.group(1).strip()
                                    distancia_min, distancia_max = 1, int(match.group(2))
                    
                    # If no distance range was found after all checks
                    if distancia_min is None:
                        clean_destino = clean_destino.split(' ')[0].strip()
                        distancia_min = 1
                        distancia_max = 1

                    # Fallback: If parsing results in an empty destination, use Origem
                    if not clean_destino:
                        clean_destino = origem
                    
                    clean_destino = clean_destino.split(' ')[0].strip()
                    clean_destino = clean_destino.split('(')[0].strip()

                    for col_idx, (vehicle, motorista) in motorista_cols.items():
                        tarifa = sheet.cell(row=row_idx, column=col_idx).value
                        if tarifa is not None and str(tarifa).strip() != "":
                            try:
                                processed_rows.append({
                                    'Transportadora': self._parse_transporter_name(file_name),
                                    'Veiculo': vehicle,
                                    'Motorista': motorista,
                                    'Origem': origem,
                                    'Destino': clean_destino,
                                    'DistanciaMin': distancia_min,
                                    'DistanciaMax': distancia_max,
                                    'Distancia': float(motorista),
                                    'Tarifa': float(tarifa),
                                    'Nomeacao': 'N/A',
                                    'Fornecedor': 'N/A',
                                    'LocalColeta': 'N/A',
                                    'Viagem': 'N/A',
                                    'Chave': f"{origem} & {clean_destino}"
                                })
                            except (ValueError, TypeError):
                                continue
                
                if processed_rows:
                    all_melted_dfs.append(pd.DataFrame(processed_rows))
            
            except Exception as e:
                print(f"  Error processing file {file_path} for 'SPOTS': {e}")
        
        return all_melted_dfs
    
    def _process_standard_fluxo(self, fluxo_path, fluxo_name):
        """Processa arquivos dos fluxos padrão (01. PRINCIPAL, 03. LINE HAUL, etc.)"""
        all_melted_dfs = []
        
        # Try to load Geoship table (search in parent and grandparent of Fluxos folder)
        geoship_df = None
        try:
            if self.tarifa_base_folder:
                # Search in parent folder (Bases)
                parent_folder = os.path.dirname(self.tarifa_base_folder)
                
                # Search in current parent
                geoship_filename = next(
                    (f for f in os.listdir(parent_folder)
                     if 'geoshiptable' in f.lower() and f.endswith(('.xlsx', '.xls'))),
                    None
                )

                # If not found, try grandparent (BC TURBO level)
                if not geoship_filename:
                    grandparent_folder = os.path.dirname(parent_folder)
                    if os.path.exists(grandparent_folder):
                        geoship_filename = next(
                            (f for f in os.listdir(grandparent_folder)
                             if 'geoshiptable' in f.lower() and f.endswith(('.xlsx', '.xls'))),
                            None
                        )
                        if geoship_filename:
                            parent_folder = grandparent_folder

                if geoship_filename:
                    geoship_full_path = os.path.join(parent_folder, geoship_filename)
                    geoship_df = pd.read_excel(geoship_full_path, engine='openpyxl')
                    geoship_df = geoship_df.rename(columns={
                        'Fornecedor': 'Fornecedor_geoship',
                        'Km Total': 'Distancia_geoship',
                        'Destino Materiais': 'Destino_geoship'
                    })
                    print(f"    ✅ Loaded GeoshipTable: '{geoship_filename}'")
                else:
                    print(f"    ℹ️  GeoshipTable not found (optional)")

        except FileNotFoundError:
            print(f"    ⚠️  Directory not found for GeoshipTable")
        except Exception as e:
            print(f"    ⚠️  Error loading GeoshipTable: {e}")
        
        # Process Excel files in fluxo folder
        for file_name in os.listdir(fluxo_path):
            if not file_name.lower().endswith(('.xlsx', '.xls')) or file_name.startswith('~'):
                continue
            file_path = os.path.join(fluxo_path, file_name)
            try:
                header_df = pd.read_excel(file_path, header=None, nrows=2, engine='openpyxl')
                header_df.iloc[0] = header_df.iloc[0].ffill()
                new_columns = []
                for i in range(len(header_df.columns)):
                    top_header = str(header_df.iloc[0, i]).upper().strip()
                    bottom_header = str(header_df.iloc[1, i]).upper().strip()
                    if 'UNNAMED' in top_header or top_header == 'NAN':
                        new_columns.append(bottom_header.lower())
                    elif 'UNNAMED' in bottom_header or bottom_header == 'NAN':
                        new_columns.append(top_header)
                    else:
                        new_columns.append(f"{top_header}_{bottom_header}")

                df = pd.read_excel(file_path, header=None, skiprows=2, engine='openpyxl')
                min_cols = min(len(df.columns), len(new_columns))
                df = df.iloc[:, :min_cols]
                df.columns = new_columns[:min_cols]

                df.columns = [col.strip() for col in df.columns]

                tipo_fluxo_col = next((col for col in df.columns if 'tipo de fluxo' in col.lower()), None)
                fornecedor_col_name = next((c for c in df.columns if 'fornecedor' in c.lower() and 'codigo' not in c.lower()), None)

                id_cols_map = {
                    'Nomeacao': next((c for c in df.columns if ('nomeação' in c.lower()) or ('nomeacao' in c.lower())), 'Nomeacao'),
                    'Origem': next((c for c in df.columns if ('cidade de coleta' in c.lower()) or ('cidade_coleta' in c.lower())), 'Origem'),
                    'LocalColeta': next((c for c in df.columns if ('local de coleta' in c.lower()) or ('local_coleta' in c.lower())), 'LocalColeta'),
                    'Destino': next((c for c in df.columns if 'destino materiais' in c.lower()), 'Destino'),
                    'Distancia': next((c for c in df.columns if 'distância' in c.lower()), 'Distancia'),
                }
                if fornecedor_col_name:
                    id_cols_map['Fornecedor'] = fornecedor_col_name

                df.rename(columns={v: k for k, v in id_cols_map.items() if v in df.columns}, inplace=True)

                if 'Fornecedor' not in df.columns:
                    df['Fornecedor'] = 'N/A'

                id_vars = list(id_cols_map.keys())
                if tipo_fluxo_col:
                    id_vars.append(tipo_fluxo_col)

                df['Transportadora'] = self._parse_transporter_name(file_name)
                id_vars.append('Transportadora')

                value_vars = [col for col in df.columns if '_' in col and col not in id_vars]

                melted_df = df.melt(
                    id_vars=[v for v in id_vars if v in df.columns],
                    value_vars=value_vars,
                    var_name='Veiculo_Viagem',
                    value_name='Tarifa'
                )

                melted_df[['Veiculo', 'Viagem']] = melted_df['Veiculo_Viagem'].str.split('_', expand=True, n=1)
                melted_df.drop('Veiculo_Viagem', axis=1, inplace=True)
                melted_df['Chave'] = melted_df['Origem'].astype(str) + ' & ' + melted_df['Destino'].astype(str)

                # Replace Geoship Rows if applicable
                if tipo_fluxo_col and tipo_fluxo_col in melted_df.columns and geoship_df is not None:
                    is_geoship = melted_df[tipo_fluxo_col].astype(str).str.lower().str.contains('geoship', na=False)
                    geoship_matches = melted_df[is_geoship]
                    non_geoship = melted_df[~is_geoship]
                    updated_rows = []

                    for _, row in geoship_matches.iterrows():
                        tipo_fluxo_value = str(row[tipo_fluxo_col]).strip()
                        geoship_key_col = next((col for col in geoship_df.columns if 'tipo' in col.lower() and 'fluxo' in col.lower()), None)
                        if geoship_key_col is None:
                            geoship_key_col = next((col for col in geoship_df.columns if 'geoship' in col.lower()), None)
                        
                        if geoship_key_col:
                            matched_geo_rows = geoship_df[
                                geoship_df[geoship_key_col].astype(str).str.lower() == tipo_fluxo_value.lower()
                            ]
                        else:
                            matched_geo_rows = pd.DataFrame()

                        if matched_geo_rows.empty:
                            updated_rows.append(row)
                        else:
                            for _, geo_row in matched_geo_rows.iterrows():
                                new_row = row.copy()
                                new_row['Fornecedor'] = geo_row.get('Fornecedor_geoship', new_row.get('Fornecedor', 'N/A'))
                                new_row['Distancia'] = geo_row.get('Distancia_geoship', new_row.get('Distancia', None))
                                new_row['Origem'] = geo_row.get('CNPJ Origem', new_row.get('Origem', None))
                                new_row['Destino'] = geo_row.get('Destino_geoship', new_row.get('Destino', None))
                                updated_rows.append(new_row)

                    melted_df = pd.concat([non_geoship, pd.DataFrame(updated_rows)], ignore_index=True).drop(columns=[tipo_fluxo_col], errors='ignore')

                all_melted_dfs.append(melted_df)

            except Exception as e:
                print(f"  Error processing file {file_path}: {e}")
                continue
        
        return all_melted_dfs
    
    def _consolidate_and_clean_data(self, all_melted_dfs, fluxo_name):
        """Consolida e limpa os DataFrames processados"""
        if not all_melted_dfs:
            return None
        
        master_df = pd.concat(all_melted_dfs, ignore_index=True).dropna(subset=['Tarifa'])
        
        # Normalize Origem
        master_df['Origem'] = master_df['Origem'].str.split(r'-').str[0].str.strip().apply(self._normalize_text)
        
        # Normalize Vehicle names
        if 'Veiculo' in master_df.columns:
            master_df['Veiculo'] = master_df['Veiculo'].apply(self._normalize_vehicle_name)
        
        # Clean string columns
        str_cols = ['Nomeacao', 'Fornecedor', 'Origem', 'LocalColeta', 'Destino', 'Veiculo', 'Viagem']
        for col in str_cols:
            if col in master_df.columns:
                master_df[col] = master_df[col].astype(str).str.strip()
        
        if 'Viagem' in master_df.columns:
            master_df['Viagem'] = master_df['Viagem'].str.upper()
            # Normalize viagem text to RT/OW codes
            viagem_map = {
                'ROUND TRIP': 'RT',
                'ROUNDTRIP': 'RT',
                'IDA E VOLTA': 'RT',
                'IDA/VOLTA': 'RT',
                'ONE WAY': 'OW',
                'ONEWAY': 'OW',
                'SOMENTE IDA': 'OW',
                'SO IDA': 'OW',
                'IDA': 'OW',
            }
            master_df['Viagem'] = master_df['Viagem'].replace(viagem_map)
        
        special_cols = ['Nomeacao', 'Fornecedor']
        for col in special_cols:
            if col in master_df.columns:
                master_df[col] = master_df[col].str.upper().str.replace(' ', '').fillna('N/A')
        
        if 'Nomeacao' in master_df.columns:
            master_df['Nomeacao'] = master_df['Nomeacao'].replace('PRINCIPALCARRETA', 'PRINCIPAL')
        
        # Convert numeric columns
        for col in ['Tarifa', 'Distancia']:
            if col in master_df.columns:
                master_df[col] = pd.to_numeric(master_df[col], errors='coerce')
        
        return master_df
    
    def _load_fluxo_from_folder(self, fluxo_path, fluxo_name):
        """Carrega dados de um fluxo específico"""
        print(f"  📁 Processing fluxo: {fluxo_name}")
        
        all_melted_dfs = []
        
        # Determine fluxo type and process accordingly
        if '04. MILK RUN' in fluxo_name or 'MILK RUN' in fluxo_name.upper():
            all_melted_dfs = self._process_milk_run_fluxo(fluxo_path, fluxo_name)
        elif 'FAIXA' in fluxo_name.upper():
            all_melted_dfs = self._process_faixa_fluxo(fluxo_path, fluxo_name)
        elif 'SPOTS' in fluxo_name.upper():
            all_melted_dfs = self._process_spots_fluxo(fluxo_path, fluxo_name)
        else:
            # Standard fluxo (01. PRINCIPAL, 03. LINE HAUL, etc.)
            all_melted_dfs = self._process_standard_fluxo(fluxo_path, fluxo_name)
        
        # Consolidate and clean
        master_df = self._consolidate_and_clean_data(all_melted_dfs, fluxo_name)
        
        if master_df is not None and not master_df.empty:
            print(f"    ✅ Loaded {len(master_df)} rows from {fluxo_name}")
            return master_df
        else:
            print(f"    ⚠️  No valid data found in {fluxo_name}")
            return None
    
    def _find_tarifa_base_folder(self):
        """Busca a pasta Fluxos dentro do database folder"""
        if not self.db_folder:
            return None
        
        db_path = Path(self.db_folder)
        
        # Look for "Fluxos" folder directly in db_folder
        fluxos_candidates = [
            db_path / "Fluxos",
            db_path / "fluxos",
            db_path / "FLUXOS",
        ]
        
        for candidate in fluxos_candidates:
            if candidate.exists() and candidate.is_dir():
                print(f"  ✓ Found Fluxos at: {candidate}")
                return str(candidate)
        
        # Look for any folder containing "fluxos" in its name
        for folder in db_path.iterdir():
            if folder.is_dir() and 'fluxos' in folder.name.lower():
                print(f"  ✓ Found Fluxos at: {folder}")
                return str(folder)
        
        return None
    
    def load_tarifa_data(self, progress_callback=None):
        """Carrega todos os dados de Tarifa (todos os fluxos)"""
        if not self.db_folder:
            print("⚠️  Database folder not set")
            return False
        
        # Find Tarifa Base folder
        self.tarifa_base_folder = self._find_tarifa_base_folder()
        
        if not self.tarifa_base_folder:
            print("ℹ️  Tarifa Base folder not found (optional)")
            return False
        
        print(f"\n{'='*60}")
        print(f"📦 Loading Tarifa Data from: {self.tarifa_base_folder}")
        print(f"{'='*60}")
        
        base_path = Path(self.tarifa_base_folder)
        
        # Get all fluxo subdirectories
        fluxos = [d for d in base_path.iterdir() if d.is_dir()]
        
        if not fluxos:
            print("⚠️  No fluxo folders found in Tarifa Base")
            return False
        
        print(f"Found {len(fluxos)} fluxo folders")
        
        # Process each fluxo
        for fluxo_dir in sorted(fluxos):
            fluxo_name = fluxo_dir.name
            
            if progress_callback:
                progress_callback(f"Loading Tarifa: {fluxo_name}...")
            
            # Check if parquet exists and is up-to-date
            parquet_path = fluxo_dir / f"{fluxo_name}.parquet"
            
            if self._needs_parquet_conversion(str(fluxo_dir), str(parquet_path)):
                # Load from Excel and create parquet
                df = self._load_fluxo_from_folder(str(fluxo_dir), fluxo_name)
                
                if df is not None and not df.empty:
                    # Save to parquet
                    try:
                        df.to_parquet(parquet_path, engine='pyarrow', compression='snappy')
                        print(f"    ✓ Parquet cache created for {fluxo_name}")
                    except Exception as e:
                        print(f"    ⚠️  Failed to create parquet: {e}")
                    
                    self.fluxo_data[fluxo_name] = df
            else:
                # Load from parquet (fast)
                try:
                    print(f"  📁 Loading {fluxo_name} from parquet cache...")
                    df = pd.read_parquet(parquet_path, engine='pyarrow')
                    self.fluxo_data[fluxo_name] = df
                    print(f"    ✅ Loaded {len(df)} rows from cache")
                except Exception as e:
                    print(f"    ⚠️  Failed to load parquet, reprocessing: {e}")
                    df = self._load_fluxo_from_folder(str(fluxo_dir), fluxo_name)
                    if df is not None:
                        self.fluxo_data[fluxo_name] = df
        
        print(f"{'='*60}")
        print(f"✅ Tarifa data loaded: {len(self.fluxo_data)} fluxos ready")
        print(f"{'='*60}\n")
        
        return True
    
    def update_db_folder(self, db_folder, progress_callback=None):
        """Atualiza o caminho do database e carrega os dados"""
        self.db_folder = db_folder
        self.fluxo_data = {}
        self.tarifa_base_folder = None
        
        return self.load_tarifa_data(progress_callback)
    
    def get_fluxo_names(self):
        """Retorna lista de nomes de fluxos disponíveis"""
        return list(self.fluxo_data.keys())
    
    def get_fluxo_data(self, fluxo_name):
        """Retorna DataFrame de um fluxo específico"""
        return self.fluxo_data.get(fluxo_name)
    
    def calculate_tariff(self, fluxo_name, origem, destino, veiculo, km_value, viagem=None):
        """
        Calcula a melhor tarifa baseada nos parâmetros fornecidos
        
        Args:
            fluxo_name: Nome do fluxo (ex: "01. PRINCIPAL")
            origem: Cidade de origem
            destino: Cidade de destino
            veiculo: Tipo de veículo
            km_value: Distância em KM
            viagem: Tipo de viagem (RT/OW) - opcional para alguns fluxos
        
        Returns:
            Dict com os resultados: {
                'status': 'success/error/not_found',
                'tarifa_original': float,
                'tarifa_real': float (calculada),
                'transportadora': str,
                'outras_opcoes': []  # Lista de outras opções disponíveis
            }
        """
        try:
            if fluxo_name not in self.fluxo_data:
                return {
                    'status': 'error',
                    'message': f'Fluxo {fluxo_name} não encontrado'
                }
            
            df = self.fluxo_data[fluxo_name].copy()
            print(f"  [Tarifa Filter] Starting with {len(df)} rows | Viagem values in data: {sorted(df['Viagem'].unique().tolist()) if 'Viagem' in df.columns else 'N/A'}")

            # Normalize viagem input to RT/OW code
            if viagem:
                _viagem_input_map = {
                    'ROUND TRIP': 'RT', 'ROUNDTRIP': 'RT', 'IDA E VOLTA': 'RT', 'IDA/VOLTA': 'RT',
                    'ONE WAY': 'OW', 'ONEWAY': 'OW', 'SOMENTE IDA': 'OW', 'SO IDA': 'OW', 'IDA': 'OW',
                }
                viagem_normalized = _viagem_input_map.get(str(viagem).strip().upper(), str(viagem).strip().upper())
            else:
                viagem_normalized = None

            # --- Filter order: Origem → Veiculo → KM range → Destino (contains) → Viagem ---

            # 1) Origem
            if origem:
                df = df[df['Origem'].str.upper() == str(origem).strip().upper()]
                print(f"  [Tarifa Filter] After Origem='{origem}': {len(df)} rows")
                # Print sample of Destino values so we can verify the format
                if 'Destino' in df.columns and not df.empty:
                    sample = df['Destino'].dropna().unique()[:15].tolist()
                    print(f"  [Tarifa Filter] Destino sample (first 15): {sample}")

            # 2) Veiculo
            if veiculo:
                df = df[df['Veiculo'].str.upper() == str(veiculo).strip().upper()]
                print(f"  [Tarifa Filter] After Veiculo='{veiculo}': {len(df)} rows")

            # 3) KM range (before Destino so we can narrow down with distance first)
            is_range_based = 'DistanciaMin' in df.columns
            df_km = df  # keep pre-km df for fallback
            if is_range_based and km_value and not df.empty:
                range_mask = (
                    pd.to_numeric(df['DistanciaMin'], errors='coerce').fillna(0) <= km_value
                ) & (
                    pd.to_numeric(df['DistanciaMax'], errors='coerce').fillna(0) >= km_value
                )
                df = df[range_mask]
                print(f"  [Tarifa Filter] After KM range ({km_value} km): {len(df)} rows")

            # 4) Destino — contains match: code like '1080' must be found inside 'FIASA(1080)'
            if destino and not df.empty:
                destino_str = str(destino).strip()
                # Strip trailing .0 from float-converted strings (e.g. '1080.0' → '1080')
                if destino_str.endswith('.0') and destino_str[:-2].isdigit():
                    destino_str = destino_str[:-2]
                df = df[df['Destino'].astype(str).str.upper().str.contains(destino_str.upper(), regex=False, na=False)]
                print(f"  [Tarifa Filter] After Destino contains '{destino_str}': {len(df)} rows")

            # 5) Viagem
            if viagem_normalized and 'Viagem' in df.columns and not df.empty:
                df = df[df['Viagem'].str.upper() == viagem_normalized]
                print(f"  [Tarifa Filter] After Viagem='{viagem_normalized}': {len(df)} rows")

            if df.empty:
                return {
                    'status': 'not_found',
                    'message': 'Nenhuma tarifa encontrada para os filtros especificados'
                }

            df_filtered = df.copy()

            # Calculate Tarifa_Real based on fluxo type
            if is_range_based:
                if 'MILK RUN' in fluxo_name.upper():
                    if km_value and km_value > 0:
                        df_filtered['Tarifa_Real'] = km_value * df_filtered['Tarifa']
                    else:
                        df_filtered['Tarifa_Real'] = df_filtered['Tarifa']  # no km → use base rate
                elif 'SPOTS' in fluxo_name.upper() and 'Distancia' in df_filtered.columns:
                    if km_value and km_value > 0:
                        df_filtered['Tarifa_Real'] = (km_value * df_filtered['Tarifa']) / df_filtered['Distancia'].replace(0, float('nan'))
                    else:
                        df_filtered['Tarifa_Real'] = df_filtered['Tarifa']
                else:  # FAIXA — tarifa is already per trip
                    df_filtered['Tarifa_Real'] = df_filtered['Tarifa']
            else:
                # Standard fluxo (e.g. Line Haul): Tarifa × KM / Distancia if km available,
                # otherwise use base tarifa directly (fixed rates indexed by Origem/Destino)
                if 'Distancia' in df_filtered.columns and km_value and km_value > 0:
                    df_filtered['Tarifa_Real'] = (km_value * df_filtered['Tarifa']) / df_filtered['Distancia'].replace(0, float('nan'))
                else:
                    df_filtered['Tarifa_Real'] = df_filtered['Tarifa']
            
            # Sort by best (cheapest) tariff
            df_sorted = df_filtered.sort_values('Tarifa_Real', ascending=True)
            
            if df_sorted.empty:
                return {
                    'status': 'not_found',
                    'message': 'Nenhuma tarifa válida calculada'
                }
            
            # Get best option
            best = df_sorted.iloc[0]
            
            # Get other options (top 5)
            others = df_sorted.iloc[1:6].to_dict('records') if len(df_sorted) > 1 else []
            
            return {
                'status': 'success',
                'tarifa_original': float(best['Tarifa']),
                'tarifa_real': float(best['Tarifa_Real']),
                'transportadora': str(best.get('Transportadora', 'N/A')),
                'veiculo': str(best.get('Veiculo', 'N/A')),
                'viagem': str(best.get('Viagem', 'N/A')),
                'origem': str(best.get('Origem', 'N/A')),
                'destino': str(best.get('Destino', 'N/A')),
                'distancia_usada': float(km_value) if km_value else 0,
                'outras_opcoes': others
            }
        
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Erro ao calcular tarifa: {str(e)}'
            }
    
    def clear_data(self):
        """Limpa todos os dados carregados"""
        self.fluxo_data = {}
        self.tarifa_base_folder = None
